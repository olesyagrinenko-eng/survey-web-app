from __future__ import annotations

import json
import os
import re
import tempfile
import traceback
import uuid
from datetime import datetime, timezone
from dataclasses import dataclass
from typing import Dict, Any, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    send_file,
    flash,
)
import math
from werkzeug.datastructures import FileStorage
from flask import after_this_request

try:
    import pyreadstat  # type: ignore
except (ImportError, OSError):  # pragma: no cover - optional at runtime
    pyreadstat = None


app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key-change-me")
# Ограничиваем размер upload, чтобы не ловить OOM/502 на очень больших файлах.
app.config["MAX_CONTENT_LENGTH"] = 40 * 1024 * 1024  # 40 MB


# --- Simple in‑memory "storage" for uploaded dataframes (per session) ---
DATASTORE: Dict[str, str] = {}  # data_key -> pickle path
OPTIONSTORE: Dict[str, Dict[str, List[Any]]] = {}  # data_key -> {var: [all possible values]}
COLUMN_LABELSTORE: Dict[str, Dict[str, str]] = {}  # data_key -> {column_name: human label / question text}
EXCELSTORE: Dict[str, str] = {}  # excel_id -> file path
RESULTSTORE: Dict[str, Dict[str, Any]] = {}  # result_id -> {"table_html": str, "significance": dict}


@dataclass
class SurveyConfig:
    row_vars: List[str]
    col_vars: List[str]
    metric_count: bool
    metric_percent: bool
    percent_base_total_sample: bool
    use_weights: bool
    perform_ztest: bool
    show_sig_marks: bool = True
    use_nested_columns: bool = False
    # Строки «ответивших / в сегменте» по каждому вопросу в строках (для ЦГ с неполным охватом)
    show_segment_coverage: bool = False


def _column_manifest_path(data_key: str) -> str:
    """Лёгкий список колонок без чтения всего pickle (страница меток после upload)."""
    return f"/tmp/survey_cols_{data_key}.json"


def _write_column_manifest(data_key: str, df: pd.DataFrame) -> None:
    payload = {
        "columns": [str(c) for c in df.columns],
        "n_rows": int(len(df)),
    }
    path = _column_manifest_path(data_key)
    with open(path, "w", encoding="utf-8") as mf:
        json.dump(payload, mf, ensure_ascii=False)


def _read_column_manifest(data_key: str) -> List[str] | None:
    path = _column_manifest_path(data_key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        cols = data.get("columns")
        if not isinstance(cols, list):
            return None
        return [str(c) for c in cols]
    except Exception:
        return None


def _filter_labels_for_df(labels: Dict[str, str], df: pd.DataFrame) -> Dict[str, str]:
    cols = {str(c) for c in df.columns}
    return {str(k): str(v).strip() for k, v in labels.items() if str(k) in cols and str(v).strip()}


def _store_dataframe(
    df: pd.DataFrame,
    all_values_map: Dict[str, List[Any]] | None = None,
    column_labels: Dict[str, str] | None = None,
) -> str:
    """Сохраняет dataframe на диск и возвращает data_key."""
    data_key = uuid.uuid4().hex
    path = f"/tmp/survey_data_{data_key}.pkl"
    df = df.copy()
    df.columns = [str(c) for c in df.columns]
    try:
        df.to_pickle(path)
    except Exception:
        # Некоторые типы из SPSS редко ломают pickle — пробуем явный protocol=4
        import pickle as _pickle

        with open(path, "wb") as pf:
            _pickle.dump(df, pf, protocol=4)
    DATASTORE[data_key] = path
    OPTIONSTORE[data_key] = all_values_map or {}
    COLUMN_LABELSTORE[data_key] = _filter_labels_for_df(column_labels or {}, df)
    try:
        _write_column_manifest(data_key, df)
    except OSError:
        pass
    return data_key


def _load_dataframe(data_key: str) -> pd.DataFrame:
    path = DATASTORE.get(data_key)
    if not path or not os.path.exists(path):
        raise ValueError("Данные не найдены во временном хранилище")
    return pd.read_pickle(path)


def _drop_dataframe(data_key: str | None) -> None:
    if not data_key:
        return
    path = DATASTORE.pop(data_key, None)
    OPTIONSTORE.pop(data_key, None)
    COLUMN_LABELSTORE.pop(data_key, None)
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass
    mp = _column_manifest_path(data_key)
    if os.path.exists(mp):
        try:
            os.remove(mp)
        except OSError:
            pass


def _clear_outputs(session_obj: Any) -> None:
    excel_id = session_obj.get("last_excel_id")
    if excel_id:
        excel_path = EXCELSTORE.pop(excel_id, None)
        if excel_path and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
            except OSError:
                pass
        session_obj.pop("last_excel_id", None)

    result_id = session_obj.get("last_result_id")
    if result_id:
        RESULTSTORE.pop(result_id, None)
        session_obj.pop("last_result_id", None)


def _replace_active_dataframe(session_obj: Any, new_df: pd.DataFrame) -> str:
    """Заменяет активную базу новой версией и чистит старую + outputs."""
    old_key = session_obj.get("active_data_key")
    inherited_options = OPTIONSTORE.get(old_key, {}).copy() if old_key else {}
    inherited_labels = COLUMN_LABELSTORE.get(old_key, {}).copy() if old_key else {}
    new_key = _store_dataframe(new_df, inherited_options, inherited_labels)
    session_obj["active_data_key"] = new_key
    if not session_obj.get("data_key"):
        session_obj["data_key"] = new_key
    _drop_dataframe(old_key)
    _clear_outputs(session_obj)
    return new_key


def _extract_all_values_from_sav_meta(meta: Any) -> Dict[str, List[Any]]:
    """
    Пытается достать полный список кодов ответов из метаданных .sav
    (включая варианты с нулевыми наблюдениями).
    """
    out: Dict[str, List[Any]] = {}
    if meta is None:
        return out

    var_value_labels = getattr(meta, "variable_value_labels", None) or {}
    if isinstance(var_value_labels, dict):
        for var_name, labels_map in var_value_labels.items():
            if isinstance(labels_map, dict):
                out[str(var_name)] = _safe_sort(list(labels_map.keys()))
        if out:
            return out

    var_to_label = getattr(meta, "variable_to_label", None) or {}
    value_labels = getattr(meta, "value_labels", None) or {}
    if isinstance(var_to_label, dict) and isinstance(value_labels, dict):
        for var_name, label_set_name in var_to_label.items():
            labels_map = value_labels.get(label_set_name)
            if isinstance(labels_map, dict):
                out[str(var_name)] = _safe_sort(list(labels_map.keys()))
    return out


def _extract_column_labels_from_meta(meta: Any) -> Dict[str, str]:
    """Текстовые метки переменных из метаданных SPSS (если они есть в .sav)."""
    if meta is None:
        return {}
    out: Dict[str, str] = {}
    d = getattr(meta, "column_names_to_labels", None)
    if isinstance(d, dict):
        for k, v in d.items():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = str(v).strip()
            if s:
                out[str(k)] = s
        if out:
            return out
    names = getattr(meta, "column_names", None)
    labels = getattr(meta, "column_labels", None)
    if isinstance(names, list) and isinstance(labels, list):
        for n, lab in zip(names, labels):
            if lab is None or (isinstance(lab, float) and pd.isna(lab)):
                continue
            s = str(lab).strip()
            if s:
                out[str(n)] = s
    return out


def _get_series_all_values(df: pd.DataFrame, var_name: str, all_values_map: Dict[str, List[Any]]) -> List[Any]:
    mapped_values = all_values_map.get(var_name, [])
    if mapped_values:
        return mapped_values

    series = df[var_name]
    if pd.api.types.is_categorical_dtype(series):
        return _safe_sort(series.cat.categories.tolist())
    return _safe_sort(series.dropna().unique())


_SPSS_SUFFIXES = (".sav", ".zsav", ".por")


def _upload_extension(filename: str) -> str:
    """
    Надёжное расширение для multipart-имени файла (в т.ч. name.something.sav).
    """
    base = os.path.basename((filename or "").strip())
    base_lower = base.lower()
    for suf in _SPSS_SUFFIXES:
        if base_lower.endswith(suf):
            return suf[1:]  # без точки
    _, ext = os.path.splitext(base_lower)
    return ext[1:] if ext.startswith(".") else ext


def _looks_like_spss_system_file(path: str) -> bool:
    """Эвристика: классический SPSS .sav начинается с сигнатуры $FL2."""
    try:
        with open(path, "rb") as f:
            head = f.read(4)
        return head == b"$FL2"
    except OSError:
        return False


def _read_spss_from_path(path: str) -> Tuple[pd.DataFrame, Any]:
    """
    Читает SPSS (.sav, .zsav через read_sav) или Portable (.por через read_por).
    Пробует несколько кодировок — частая причина сбоев на русских метках.
    """
    if pyreadstat is None:
        raise RuntimeError(
            "Чтение SPSS недоступно: не установлен пакет pyreadstat. "
            "На сервере выполните установку зависимостей из requirements.txt "
            "(pip install -r requirements.txt)."
        )

    ext = os.path.splitext(path)[1].lower()
    encodings: List[Any] = [None, "utf-8", "latin-1", "cp1251", "cp1252"]

    def _try_read_por(enc: Any) -> Tuple[pd.DataFrame, Any]:
        read_por = getattr(pyreadstat, "read_por", None)
        if read_por is None:
            raise RuntimeError("Формат .por не поддерживается этой версией pyreadstat.")
        if enc is None:
            return read_por(path)
        try:
            return read_por(path, encoding=enc)
        except TypeError:
            return read_por(path)

    def _try_read_sav(enc: Any) -> Tuple[pd.DataFrame, Any]:
        if enc is None:
            return pyreadstat.read_sav(path)
        try:
            return pyreadstat.read_sav(path, encoding=enc)
        except TypeError:
            return pyreadstat.read_sav(path)

    last_err: Exception | None = None
    for enc in encodings:
        try:
            if ext == ".por":
                df, meta = _try_read_por(enc)
            else:
                df, meta = _try_read_sav(enc)
            return df, meta
        except Exception as exc:
            last_err = exc
            continue

    raise RuntimeError(
        "Не удалось прочитать SPSS-файл (проверьте, что это не повреждённый .sav и что файл не зашифрован паролем). "
        f"Последняя ошибка: {last_err}"
    ) from last_err


def _column_name_looks_like_technical_code(c: str) -> bool:
    """Короткое имя переменной (Q1, var_01), а не формулировка вопроса в заголовке столбца."""
    s = str(c).strip()
    if not s:
        return True
    if len(s) > 72:
        return False
    if any(ch.isspace() for ch in s):
        return False
    if s.isdigit():
        return True
    if re.match(r"^[QV]\d+[a-z]?$", s, re.IGNORECASE):
        return True
    if re.match(r"^[A-Za-z][A-Za-z0-9_.]{0,70}$", s):
        return True
    return False


def _cell_looks_like_pure_survey_value(v: Any) -> bool:
    """Число или числовая строка — похоже на ответ, а не на текст метки."""
    if pd.isna(v):
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float, np.integer, np.floating)):
        return True
    s = str(v).strip().replace("\u00a0", "").replace(",", ".")
    if not s:
        return False
    if s.isdigit():
        return True
    try:
        float(s)
        return True
    except ValueError:
        return False


def _maybe_promote_first_row_as_labels(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Если первая строка данных похожа на строку текстовых меток под короткими именами столбцов
    (частый экспорт: строка заголовка — коды, следующая — формулировки), переносим её в метки и удаляем из данных.
    """
    if df.shape[0] < 2:
        return df, {}
    first = df.iloc[0]
    code_cols = [c for c in df.columns if _column_name_looks_like_technical_code(str(c))]
    if len(code_cols) < 2:
        return df, {}
    numeric_like = 0
    total_cells = 0
    for c in df.columns:
        total_cells += 1
        if _cell_looks_like_pure_survey_value(first[c]):
            numeric_like += 1
    if total_cells and (numeric_like / total_cells) > 0.38:
        return df, {}
    labels: Dict[str, str] = {}
    for c in code_cols:
        v = first[c]
        if pd.isna(v):
            continue
        vs = str(v).strip()
        if not vs:
            continue
        if _cell_looks_like_pure_survey_value(v) and len(vs) <= 12:
            continue
        c_str = str(c)
        plausible = (" " in vs) or (len(vs) >= 18) or (not vs.isascii() and len(vs) >= 10)
        if not plausible:
            continue
        if len(vs) < len(c_str) + 4 and len(vs) < 16:
            continue
        labels[c_str] = vs
    need = max(2, math.ceil(0.55 * len(code_cols)))
    if len(labels) < need:
        return df, {}
    out = df.iloc[1:].copy().reset_index(drop=True)
    return out, labels


def _load_dataframe_from_upload(file_storage) -> Tuple[pd.DataFrame, Dict[str, List[Any]], Dict[str, str]]:
    filename = file_storage.filename or ""
    ext = _upload_extension(filename)

    if ext == "csv":
        df = pd.read_csv(file_storage)
        df, promoted = _maybe_promote_first_row_as_labels(df)
        return df, {}, promoted
    if ext in {"xls", "xlsx"}:
        df = pd.read_excel(file_storage)
        df, promoted = _maybe_promote_first_row_as_labels(df)
        return df, {}, promoted
    if ext in {"sav", "zsav", "por"}:
        suffix = f".{ext}" if ext else ".sav"
        fd, tmp_path = tempfile.mkstemp(suffix=suffix)
        os.close(fd)
        try:
            try:
                file_storage.stream.seek(0)
            except (OSError, AttributeError, ValueError):
                pass
            file_storage.save(tmp_path)
            df, meta = _read_spss_from_path(tmp_path)
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        return df, _extract_all_values_from_sav_meta(meta), _extract_column_labels_from_meta(meta)

    # Расширение не распознано: пробуем определить SPSS по сигнатуре (загрузка без .sav в имени)
    fd, tmp_path = tempfile.mkstemp(suffix=".bin")
    os.close(fd)
    try:
        try:
            file_storage.stream.seek(0)
        except (OSError, AttributeError, ValueError):
            pass
        file_storage.save(tmp_path)
        if _looks_like_spss_system_file(tmp_path):
            df, meta = _read_spss_from_path(tmp_path)
            return df, _extract_all_values_from_sav_meta(meta), _extract_column_labels_from_meta(meta)
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    raise ValueError(
        "Неподдерживаемый формат файла. Разрешены: csv, xls, xlsx, sav, zsav (SPSS), por (SPSS Portable)."
    )


def _safe_sort(values: np.ndarray | List[Any]) -> List[Any]:
    try:
        numeric_values: List[Tuple[float, Any]] = []
        for v in values:
            try:
                numeric_values.append((float(v), v))
            except (TypeError, ValueError):
                # как только встречаем не‑число, сортируем всё как строки
                return sorted(values, key=lambda x: str(x))
        return [orig for _, orig in sorted(numeric_values)]
    except Exception:
        return sorted(values, key=lambda x: str(x))


def _parse_datamap_excel(path: str) -> Dict[str, str]:
    """
    Datamap в стиле Lighthouse/SSI: лист «Variable Information», строка заголовков с «Переменная» и «Метка».
    """
    xl = pd.ExcelFile(path)
    sheet: str | None = None
    for sn in xl.sheet_names:
        sl = (sn or "").lower().replace("ё", "е")
        if "variable information" in sl or "информация о переменн" in sl:
            sheet = sn
            break
    if sheet is None:
        sheet = xl.sheet_names[0]

    df: pd.DataFrame | None = None
    last_err: Exception | None = None
    for hdr in (1, 0):
        try:
            t = pd.read_excel(path, sheet_name=sheet, header=hdr)
            if t.shape[1] < 1:
                continue
            df = t
            break
        except Exception as exc:
            last_err = exc
            df = None
    if df is None:
        raise ValueError(f"Не удалось прочитать лист «{sheet}» в datamap: {last_err}")

    cols_lc = {str(c).strip().lower(): c for c in df.columns}
    code_key: str | None = None
    label_key: str | None = None
    for ck in ("переменная", "variable", "var", "name"):
        if ck in cols_lc:
            code_key = cols_lc[ck]
            break
    for lk in ("метка", "label", "question", "вопрос"):
        if lk in cols_lc:
            label_key = cols_lc[lk]
            break
    if label_key is None:
        for low, orig in cols_lc.items():
            if (("метк" in low) or low == "label") and "знач" not in low and "value" not in low:
                label_key = orig
                break
    if code_key is None:
        for low, orig in cols_lc.items():
            if "перемен" in low or low in ("variable", "name", "var"):
                code_key = orig
                break
    if code_key is None or label_key is None:
        raise ValueError(
            "В datamap не найдены колонки кода переменной и метки вопроса "
            "(ожидаются «Переменная» + «Метка» или аналоги name/label)."
        )

    out: Dict[str, str] = {}
    for _, row in df.iterrows():
        raw_code = row.get(code_key)
        if raw_code is None or (isinstance(raw_code, float) and pd.isna(raw_code)):
            continue
        code = str(raw_code).strip()
        if not code:
            continue
        raw_lab = row.get(label_key)
        if raw_lab is None or (isinstance(raw_lab, float) and pd.isna(raw_lab)):
            continue
        lab = str(raw_lab).strip()
        if not lab:
            continue
        out[code] = lab
    if not out:
        raise ValueError("В datamap не найдено ни одной пары «код переменной — метка».")
    return out


def _apply_datamap_merge(
    data_key: str,
    column_names: List[str],
    mapping: Dict[str, str],
    merge_mode: str,
    case_insensitive: bool,
) -> Tuple[int, int]:
    """(совпало кодов с колонками базы, записано/обновлено меток)."""
    canon = [str(c) for c in column_names]
    canon_set = set(canon)
    lower_map = {c.lower(): c for c in canon} if case_insensitive else {}
    cur = COLUMN_LABELSTORE.setdefault(data_key, {})
    matched = 0
    updated = 0
    for raw_k, raw_v in mapping.items():
        k = str(raw_k).strip()
        if not k:
            continue
        if raw_v is None or (isinstance(raw_v, float) and pd.isna(raw_v)):
            continue
        v = str(raw_v).strip()
        if not v:
            continue
        target = k if k in canon_set else None
        if target is None and case_insensitive:
            target = lower_map.get(k.lower())
        if target is None:
            continue
        matched += 1
        if merge_mode == "fill_empty":
            existing = str(cur.get(target, "")).strip()
            if existing:
                continue
        cur[target] = v
        updated += 1
    return matched, updated


def _ui_column_items_from_columns(columns: List[str], data_key: str) -> List[Dict[str, Any]]:
    labels = COLUMN_LABELSTORE.get(data_key, {})
    out: List[Dict[str, Any]] = []
    for col in columns:
        code = str(col)
        lab = str(labels.get(code, "")).strip()
        header_descriptive = not _column_name_looks_like_technical_code(code)
        out.append(
            {
                "code": code,
                "label": lab,
                "has_label": bool(lab),
                "header_is_descriptive": header_descriptive,
            }
        )
    # Порядок как в исходной таблице (колонки df / manifest), не по алфавиту меток.
    return out


def _ui_column_items(df: pd.DataFrame, data_key: str) -> List[Dict[str, Any]]:
    return _ui_column_items_from_columns([str(c) for c in df.columns], data_key)


def _ui_column_groups(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    from collections import defaultdict

    buckets: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    first_pos: Dict[str, int] = {}
    for pos, it in enumerate(items):
        code = it["code"]
        if "_" in code:
            root = code.split("_", 1)[0]
        else:
            m = re.match(r"^([A-Za-zА-Яа-яЁё]+)", code)
            root = m.group(1) if m else code[:1]
        if root not in first_pos:
            first_pos[root] = pos
        buckets[root].append(it)
    # Группы — по первому появлению префикса в базе; внутри группы — порядок из базы.
    ordered_roots = sorted(first_pos.keys(), key=lambda r: (first_pos[r], r.lower()))
    groups: List[Dict[str, Any]] = []
    for root in ordered_roots:
        groups.append(
            {
                "name": root,
                # ключ не «items»: в Jinja group.items — это метод dict, а не список колонок
                "members": buckets[root],
            }
        )
    return groups


def _template_column_bundle(df: pd.DataFrame, data_key: str) -> Dict[str, Any]:
    items = _ui_column_items(df, data_key)
    return {
        "columns": list(df.columns),
        "column_items": items,
        "column_groups": _ui_column_groups(items),
        "n_column_labels": sum(1 for x in items if x["has_label"] or x["header_is_descriptive"]),
    }


# Две ведущие колонки кросс-таблицы: метка вопроса (datamap / код) и категория ответа.
CROSSTAB_ROW_LABEL_Q = "Метка вопроса"
CROSSTAB_ROW_LABEL_A = "Текстовая формулировка"


def _crosstab_category_text(rv: Any) -> str:
    """Текст во второй колонке строки (категория / значение показателя)."""
    if pd.isna(rv):
        return ""
    try:
        if isinstance(rv, (float, np.floating)) and float(rv) == int(float(rv)):
            return str(int(float(rv)))
    except (ValueError, TypeError, OverflowError):
        pass
    if isinstance(rv, (int, np.integer)):
        return str(int(rv))
    return str(rv)


def _significance_key_serialize(row_q: str, row_a: str, col_label: str) -> str:
    """Стабильный ключ для JSON/hранилища (без разделителя || внутри полей)."""
    return json.dumps([row_q, row_a, col_label], ensure_ascii=False, separators=(",", ":"))


def _crosstab_cell_str(val: Any) -> str:
    """Строка для сопоставления строк таблицы (пусто вместо nan/None)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none"):
        return ""
    return s


def _significance_dict_from_storage(raw: Dict[str, Any]) -> Dict[Tuple[str, str, str], Dict[str, Any]]:
    """Восстанавливает словарь значимостей из RESULTSTORE (новый и старый формат ключей)."""
    out: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for k, v in raw.items():
        if not isinstance(k, str):
            continue
        try:
            a = json.loads(k)
            if isinstance(a, list) and len(a) == 3:
                out[(str(a[0]), str(a[1]), str(a[2]))] = v
                continue
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
        if "||" in k:
            row_part, ck = k.split("||", 1)
            out[(row_part, "", ck)] = v
        else:
            out[(k, "", "")] = v
    return out


def _z_test_vs_total(
    total_success: float,
    total_n: float,
    group_success: float,
    group_n: float,
) -> Tuple[float, float]:
    """
    Классический пропорционный z‑тест (две независимые выборки).
    Возвращает (z, p_value).
    """
    if total_n <= 0 or group_n <= 0:
        return 0.0, 1.0

    p1 = group_success / group_n
    p2 = total_success / total_n
    p_pool = (group_success + total_success) / (group_n + total_n)

    denom = np.sqrt(p_pool * (1 - p_pool) * (1 / group_n + 1 / total_n))
    if denom == 0:
        return 0.0, 1.0

    z = (p1 - p2) / denom
    # нормальное CDF через erf, чтобы не тянуть scipy
    p = 2 * (1 - 0.5 * (1 + math.erf(abs(z) / math.sqrt(2))))
    return float(z), float(p)


def build_crosstab(
    df: pd.DataFrame,
    config: SurveyConfig,
    all_values_map: Dict[str, List[Any]] | None = None,
    column_labels: Dict[str, str] | None = None,
) -> Tuple[pd.DataFrame, Dict[Tuple[str, str, str], Dict[str, Any]]]:
    """
    Строит кросс‑таблицу.
    Возвращает:
      - таблицу значений (две ведущие колонки: метка вопроса и текстовая формулировка категории)
      - словарь значимостей {(метка_вопроса, формулировка, заголовок_столбца) -> {"z", "p", "direction"}}
    column_labels: подписи переменных (.sav / datamap) для меток вопросов и заголовков сегментов.
    """
    weight_col = "weight" if config.use_weights and "weight" in df.columns else None
    all_values_map = all_values_map or {}
    c_labels = {
        str(k): str(v).strip()
        for k, v in (column_labels or {}).items()
        if str(v).strip()
    }

    def _var_display(code: str) -> str:
        c = str(code)
        return c_labels[c] if c in c_labels else c

    # Ограничения, чтобы не падать по памяти на Render
    MAX_ROW_UNIQUE_SKIP = 100
    MAX_ROW_VALUES = 50
    MAX_COL_VALUES = 50
    MAX_TOTAL_CELLS = 50000  # оценка до фактического построения

    # Оценка размера результата (сколько строк * сколько столбцов)
    est_row_count = 0
    for row_var in config.row_vars:
        uniq = _get_series_all_values(df, row_var, all_values_map)
        n_uniq = len(uniq)
        if n_uniq > MAX_ROW_UNIQUE_SKIP:
            continue
        est_row_count += min(n_uniq, MAX_ROW_VALUES)

    est_col_values = 0
    for col_var in config.col_vars:
        uniq = _get_series_all_values(df, col_var, all_values_map)
        est_col_values += min(len(uniq), MAX_COL_VALUES)

    est_total_cols = 2 + est_col_values  # две колонки описания строки + Total и сегменты
    cov_extra = len(config.row_vars) if config.show_segment_coverage else 0
    est_cells = (est_row_count + 1 + cov_extra) * est_total_cols
    if est_row_count <= 0:
        raise ValueError("Не удалось построить таблицу: слишком много уникальных значений/нет данных.")
    if est_cells > MAX_TOTAL_CELLS:
        raise ValueError(
            f"Слишком большой объём кросс-таблицы (оценка ячеек: {est_cells}). "
            f"Уменьшите количество выбранных переменных или снимите 'Выбрать все'. "
            f"(Ограничение: {MAX_TOTAL_CELLS} ячеек)"
        )

    # Подготовка столбцов анализа.
    # По умолчанию: одноуровневые столбцы, даже если выбрано несколько переменных.
    # Комбинации (второй уровень) строим только по явному флажку.
    analysis_defs: List[Dict[str, Any]] = [{"label": "Total", "mask": None}]
    if len(config.col_vars) == 1 or not config.use_nested_columns:
        # Базовое поведение: раскрываем каждую выбранную переменную отдельно
        for col_var in config.col_vars:
            uniques = _get_series_all_values(df, col_var, all_values_map)
            if len(uniques) > MAX_COL_VALUES:
                uniques = uniques[:MAX_COL_VALUES]
            for val in uniques:
                mask = df[col_var] == val
                analysis_defs.append({"label": f"{_var_display(col_var)}: {val}", "mask": mask})
    else:
        # Опциональный второй уровень: комбинации выбранных столбцов
        cols_df = df[config.col_vars].dropna()
        combo_rows = cols_df.drop_duplicates()
        combo_rows = combo_rows.assign(
            __sort_key__=combo_rows.apply(lambda r: " | ".join([str(r[c]) for c in config.col_vars]), axis=1)
        ).sort_values("__sort_key__")
        if len(combo_rows) > MAX_COL_VALUES:
            combo_rows = combo_rows.head(MAX_COL_VALUES)
        for _, combo in combo_rows.iterrows():
            parts = []
            mask = pd.Series(True, index=df.index)
            for c in config.col_vars:
                val = combo[c]
                parts.append(f"{_var_display(c)}: {val}")
                mask = mask & (df[c] == val)
            analysis_defs.append({"label": " | ".join(parts), "mask": mask})

    # заголовки
    headers = [CROSSTAB_ROW_LABEL_Q, CROSSTAB_ROW_LABEL_A]
    headers.extend([d["label"] for d in analysis_defs])

    rows: List[List[Any]] = []
    significance: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    # первая строка — невзвешенные размеры выборки
    sample_row = ["Выборка (N)", ""]
    for d in analysis_defs:
        mask = d["mask"]
        n = len(df) if mask is None else int(mask.sum())
        sample_row.append(n)
    rows.append(sample_row)

    if config.show_segment_coverage:
        for row_var in config.row_vars:
            if row_var not in df.columns:
                continue
            answered = df[row_var].notna()
            cov_row: List[Any] = [
                _var_display(row_var),
                "Охват по ЦГ (ответивших / в сегменте)",
            ]
            for d in analysis_defs:
                col_mask = d["mask"]
                seg = pd.Series(True, index=df.index) if col_mask is None else col_mask
                # Как строка «Выборка (N)»: число кейсов, без весов (сопоставимо с долей «не все ответили»).
                n_seg = int(seg.sum())
                n_answ = int((seg & answered).sum())
                cov_row.append(f"{n_answ} / {n_seg}")
            rows.append(cov_row)

    for row_var in config.row_vars:
        unique_vals = _get_series_all_values(df, row_var, all_values_map)
        if len(unique_vals) > MAX_ROW_UNIQUE_SKIP:
            continue
        limited_row_values = unique_vals[:MAX_ROW_VALUES] if len(unique_vals) > MAX_ROW_VALUES else unique_vals
        for rv in limited_row_values:
            row_q = _var_display(row_var)
            row_a = _crosstab_category_text(rv)
            row: List[Any] = [row_q, row_a]

            # для z‑теста нам нужны success и n для total и каждой группы (по percent)
            total_success = 0.0
            total_base = 0.0
            group_success_list: List[Tuple[int, float, float]] = []  # index, success, n

            for idx, d in enumerate(analysis_defs):
                col_mask = d["mask"]
                if col_mask is None:
                    # База процента: от ответивших (default) или от всей выборки.
                    base_mask = pd.Series(True, index=df.index) if config.percent_base_total_sample else df[row_var].notna()
                    success_mask = (df[row_var] == rv) & base_mask
                else:
                    base_mask = col_mask if config.percent_base_total_sample else (col_mask & df[row_var].notna())
                    success_mask = (df[row_var] == rv) & col_mask

                if weight_col:
                    base_n = df.loc[base_mask, weight_col].sum()
                    success_n = df.loc[success_mask, weight_col].sum()
                else:
                    base_n = base_mask.sum()
                    success_n = success_mask.sum()

                # count
                if config.metric_count:
                    value = success_n
                else:
                    value = success_n

                # percent
                if config.metric_percent:
                    if base_n > 0:
                        value = round((success_n / base_n) * 100, 1)
                    else:
                        value = 0.0

                row.append(value)

                # подготовка данных для z‑теста
                if config.perform_ztest and config.metric_percent and base_n > 0:
                    if idx == 0:
                        total_success = success_n
                        total_base = base_n
                    else:
                        group_success_list.append((idx, success_n, base_n))

            # после прохода по столбцам считаем z‑тесты
            if config.perform_ztest and config.metric_percent and total_base > 0:
                for idx, g_succ, g_n in group_success_list:
                    z, p = _z_test_vs_total(
                        total_success=total_success,
                        total_n=total_base,
                        group_success=g_succ,
                        group_n=g_n,
                    )
                    direction = "none"
                    if p < 0.05:
                        direction = "up" if z > 0 else "down"
                    col_label = analysis_defs[idx]["label"]
                    significance[(row_q, row_a, col_label)] = {
                        "z": z,
                        "p": p,
                        "direction": direction,
                    }

            rows.append(row)

    table_df = pd.DataFrame(rows, columns=headers)
    return table_df, significance


def _annotate_table_with_significance(
    table_df: pd.DataFrame,
    significance: Dict[Tuple[str, str, str], Dict[str, Any]],
) -> pd.DataFrame:
    """
    Добавляет маркеры значимости прямо в значения таблицы:
    - ↑ если p<0.05 и выше Total
    - ↓ если p<0.05 и ниже Total
    """
    if table_df.empty or not significance:
        return table_df.copy()

    out = table_df.copy()
    if CROSSTAB_ROW_LABEL_Q not in out.columns or CROSSTAB_ROW_LABEL_A not in out.columns:
        return out

    row_index: Dict[Tuple[str, str], int] = {}
    iq = out.columns.get_loc(CROSSTAB_ROW_LABEL_Q)
    ia = out.columns.get_loc(CROSSTAB_ROW_LABEL_A)
    for idx in range(len(out)):
        rq = _crosstab_cell_str(out.iat[idx, iq])
        ra = _crosstab_cell_str(out.iat[idx, ia])
        row_index[(rq, ra)] = idx

    for (row_q, row_a, col_label), info in significance.items():
        if col_label not in out.columns:
            continue
        r_idx = row_index.get((_crosstab_cell_str(row_q), _crosstab_cell_str(row_a)))
        if r_idx is None:
            continue

        direction = str(info.get("direction", "none"))
        if direction not in {"up", "down"}:
            continue

        current = out.at[r_idx, col_label]
        arrow = "↑" if direction == "up" else "↓"
        out.at[r_idx, col_label] = f"{current} {arrow}"

    return out


def _apply_significance_fill_to_excel(
    excel_path: str,
    table_df: pd.DataFrame,
    significance: Dict[Tuple[str, str, str], Dict[str, Any]],
) -> None:
    """
    Красит ячейки значимых отличий в Excel (без изменения значения ячейки).
    up   -> зеленая заливка
    down -> красная заливка
    """
    if not significance:
        return
    wb = load_workbook(excel_path)
    ws = wb.active

    header_to_col_idx: Dict[str, int] = {}
    for idx, col_name in enumerate(table_df.columns.tolist(), start=1):
        header_to_col_idx[str(col_name)] = idx

    rowlabel_to_row_idx: Dict[Tuple[str, str], int] = {}
    if CROSSTAB_ROW_LABEL_Q in table_df.columns and CROSSTAB_ROW_LABEL_A in table_df.columns:
        iq = table_df.columns.get_loc(CROSSTAB_ROW_LABEL_Q)
        ia = table_df.columns.get_loc(CROSSTAB_ROW_LABEL_A)
        for i in range(len(table_df)):
            rq = _crosstab_cell_str(table_df.iat[i, iq])
            ra = _crosstab_cell_str(table_df.iat[i, ia])
            rowlabel_to_row_idx[(rq, ra)] = i + 2
    else:
        # Старый формат с одной колонкой «Переменная»
        legacy = "Переменная"
        if legacy in table_df.columns:
            il = table_df.columns.get_loc(legacy)
            for i in range(len(table_df)):
                rowlabel_to_row_idx[(_crosstab_cell_str(table_df.iat[i, il]), "")] = i + 2

    fill_up = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fill_down = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for (row_q, row_a, col_label), info in significance.items():
        direction = str(info.get("direction", "none"))
        if direction not in {"up", "down"}:
            continue
        excel_row = rowlabel_to_row_idx.get(
            (_crosstab_cell_str(row_q), _crosstab_cell_str(row_a))
        )
        if excel_row is None:
            excel_row = rowlabel_to_row_idx.get((_crosstab_cell_str(row_q), ""))
        excel_col = header_to_col_idx.get(str(col_label))
        if excel_row is None or excel_col is None:
            continue
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.fill = fill_up if direction == "up" else fill_down

    wb.save(excel_path)


def _is_probably_open_question(series: pd.Series) -> bool:
    """
    Эвристика: считаем открытым вопросом текстовую переменную
    с высокой уникальностью и/или длинными неструктурированными ответами.
    """
    s = series.dropna().astype(str)
    if len(s) < 20:
        return False
    unique_ratio = s.nunique() / max(len(s), 1)
    avg_len = s.str.len().mean() if len(s) else 0
    # если очень много уникальных и ответы длинные -> open
    if unique_ratio > 0.7 and avg_len > 12:
        return True
    # или просто очень длинные ответы
    if avg_len > 25:
        return True
    return False


def _detect_closed_columns(df: pd.DataFrame) -> List[str]:
    closed: List[str] = []
    for c in df.columns:
        series = df[c]
        # numeric/bool чаще всего закрытые
        if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_bool_dtype(series):
            closed.append(c)
            continue
        # object: отсекаем "open text"
        if not _is_probably_open_question(series):
            closed.append(c)
    return closed


# =========================
# Weighting (bot logic)
# =========================

TRIM_MIN = 0.3
TRIM_MAX = 3.0
MAX_VARS = 10
MAX_CATS_PER_DIM = 60  # for safety (as in the telegram bot)
NORMALIZE_MEAN_1 = True


def create_cross_variable(df: pd.DataFrame, vars_list: List[str]) -> pd.Series:
    """Creates cross variable as in telegram weighting bot: 'A × B × C'."""
    if len(vars_list) == 1:
        return df[vars_list[0]].astype(str)
    cross_var = df[vars_list[0]].astype(str)
    for var in vars_list[1:]:
        cross_var = cross_var + " × " + df[var].astype(str)
    return cross_var


def parse_targets_input(text: str, categories: List[str]) -> Dict[str, float]:
    """
    Parse text like:
      - "M=48, F=52" (percent)
      - "M:0.48, F:0.52" (shares)
    Returns normalized dict {category: share(0..1)} with sum=1.
    """
    raw = (text or "").strip()
    if not raw:
        raise ValueError("пустой ввод")

    norm = raw.replace("\n", " ")
    pairs = re.split(r"[;,|]", norm)
    out: Dict[str, float] = {}
    cats_lower = [c.lower() for c in categories]

    def to_float(v: str) -> float:
        v = v.strip().replace("%", "")
        v = v.replace(",", ".")
        try:
            return float(v)
        except Exception:
            raise ValueError(f"некорректное число: '{v}'")

    # If it's just numbers separated by delimiters (no explicit keys)
    if all("=" not in p and ":" not in p for p in pairs):
        nums = [to_float(p) for p in pairs if p.strip()]
        if len(nums) != len(categories):
            raise ValueError(
                f"ожидалось {len(categories)} чисел (для категорий: {', '.join(categories)}), получили {len(nums)}"
            )
        s = sum(nums)
        vals = [n / 100.0 if 95 <= s <= 105 else n for n in nums]
        s2 = sum(vals)
        if s2 <= 0:
            raise ValueError("сумма долей должна быть > 0")
        vals = [v / s2 for v in vals]
        return {cat: vals[i] for i, cat in enumerate(categories)}

    # key=value or key:value
    for p in pairs:
        if not p.strip():
            continue
        if "=" in p:
            key, val = p.split("=", 1)
        elif ":" in p:
            key, val = p.split(":", 1)
        else:
            raise ValueError(f"не распознан фрагмент: '{p.strip()}' (ожидается cat=val)")

        key = key.strip()
        valf = to_float(val)

        if key.lower() in cats_lower:
            cat = categories[cats_lower.index(key.lower())]
        else:
            if key not in categories:
                raise ValueError(f"категория '{key}' отсутствует. Доступные: {', '.join(categories)}")
            cat = key

        out[cat] = valf

    s = sum(out.values())
    # If it looks like percent
    if 95 <= s <= 105:
        out = {k: v / 100.0 for k, v in out.items()}
        s = sum(out.values())

    if set(out.keys()) != set(categories):
        missing = [c for c in categories if c not in out]
        extra = [c for c in out if c not in categories]
        msg = []
        if missing:
            msg.append("нет значений для: " + ", ".join(missing))
        if extra:
            msg.append("лишние категории: " + ", ".join(extra))
        raise ValueError("; ".join(msg))

    if s <= 0:
        raise ValueError("сумма долей должна быть > 0")

    out = {k: v / s for k, v in out.items()}
    return out


def read_targets_from_excel(file_storage: FileStorage) -> Dict[str, float]:
    df = pd.read_excel(file_storage)

    category_col = None
    target_col = None

    category_names = ["category", "категория", "cat", "name", "название", "группа", "group"]
    target_names = [
        "target",
        "targets",
        "таргет",
        "таргеты",
        "цель",
        "доля",
        "share",
        "percent",
        "процент",
        "value",
        "значение",
    ]

    for col in df.columns:
        if str(col).lower().strip() in category_names:
            category_col = col
            break

    for col in df.columns:
        if str(col).lower().strip() in target_names:
            target_col = col
            break

    if category_col is None or target_col is None:
        cols = list(df.columns)
        if len(cols) >= 2:
            category_col = cols[0]
            target_col = cols[1]
        else:
            raise ValueError(f"Недостаточно колонок в файле. Найдено: {cols}")

    targets: Dict[str, float] = {}
    for _, row in df.iterrows():
        cat = str(row[category_col]).strip()
        try:
            target = float(row[target_col])
        except (ValueError, TypeError):
            continue
        targets[cat] = target

    if not targets:
        raise ValueError("Не удалось прочитать ни одной пары категория-значение")

    total = sum(targets.values())
    if total > 0:
        targets = {k: v / total for k, v in targets.items()}

    return targets


def create_targets_template_xlsx(categories: List[str], current_shares: pd.Series, out_path: str) -> None:
    template_data = []
    for category in sorted(categories):
        template_data.append(
            {
                "Category": category,
                "Targets": float(current_shares.get(category, 0.0)),
            }
        )
    template_df = pd.DataFrame(template_data)
    template_df.to_excel(out_path, index=False, engine="openpyxl")


def clean_sheet_name(name: str) -> str:
    """Очистка названия листа Excel."""
    invalid_chars = r"[\\/:*?\[\]]"
    clean_name = re.sub(invalid_chars, "_", str(name))
    clean_name = clean_name.strip().rstrip(".")
    clean_name = clean_name[:31]
    return clean_name or "Sheet"


def create_sequential_targets_template_xlsx(
    df: pd.DataFrame,
    vars_list: List[str],
    out_path: str,
) -> None:
    """
    Делает Excel workbook: по листу на каждую переменную.
    Каждый лист: Category / Targets (текущие доли как стартовые значения).
    """
    import pandas as _pd

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for var in vars_list:
            series = df[var].dropna().astype(str)
            categories = sorted(series.unique().tolist())
            current_shares = series.value_counts(normalize=True).sort_index()
            template_data = []
            for cat in categories:
                template_data.append({"Category": cat, "Targets": float(current_shares.get(cat, 0.0))})
            template_df = _pd.DataFrame(template_data)
            writer.sheets  # touch to ensure writer init
            template_df.to_excel(writer, sheet_name=clean_sheet_name(var), index=False)


def _read_targets_from_dataframe(df_targets: pd.DataFrame) -> Dict[str, float]:
    category_col = None
    target_col = None

    category_names = ["category", "категория", "cat", "name", "название", "группа", "group"]
    target_names = [
        "target",
        "targets",
        "таргет",
        "таргеты",
        "цель",
        "доля",
        "share",
        "percent",
        "процент",
        "value",
        "значение",
    ]

    for col in df_targets.columns:
        if str(col).lower().strip() in category_names:
            category_col = col
            break
    for col in df_targets.columns:
        if str(col).lower().strip() in target_names:
            target_col = col
            break

    if category_col is None or target_col is None:
        cols = list(df_targets.columns)
        if len(cols) >= 2:
            category_col = cols[0]
            target_col = cols[1]
        else:
            raise ValueError(f"Недостаточно колонок в sheet: {cols}")

    targets: Dict[str, float] = {}
    for _, row in df_targets.iterrows():
        cat = str(row[category_col]).strip()
        try:
            target = float(row[target_col])
        except (ValueError, TypeError):
            continue
        targets[cat] = target

    if not targets:
        raise ValueError("Не удалось прочитать ни одной пары категория-значение на sheet")

    total = sum(targets.values())
    if 95 <= total <= 105:
        # похоже на проценты
        targets = {k: v / 100.0 for k, v in targets.items()}
    else:
        targets = {k: v / total for k, v in targets.items()} if total > 0 else targets

    return targets


def read_sequential_targets_from_excel(file_storage: FileStorage, vars_list: List[str]) -> Dict[str, Dict[str, float]]:
    workbook = pd.read_excel(file_storage, sheet_name=None)

    out: Dict[str, Dict[str, float]] = {}
    for var in vars_list:
        sheet = clean_sheet_name(var)
        if sheet not in workbook:
            raise ValueError(f"В файле нет sheet '{sheet}' для переменной '{var}'")
        df_sheet = workbook[sheet]
        out[var] = _read_targets_from_dataframe(df_sheet)

    # Validate that each sheet contains exactly the same set of categories as in data
    # (слабая проверка: гарантируем, что хотя бы все заданные категории есть в данных).
    return out


def sequential_weight(df: pd.DataFrame, vars_list: List[str], targets_by_var: Dict[str, Dict[str, float]]) -> pd.DataFrame:
    """
    Реплика логики sequential-взвешивания из Telegram-бота:
    веса корректируются по очереди переменных, используя unweighted value_counts(normalize=True).
    """
    weights = pd.Series(1.0, index=df.index)

    for var in vars_list:
        # категории в данных (как строки)
        current_shares = df[var].dropna().astype(str).value_counts(normalize=True)
        if var not in targets_by_var:
            raise ValueError(f"Нет targets для переменной '{var}'")

        targets = targets_by_var[var]

        for category, target_share in targets.items():
            current_share = float(current_shares.get(str(category), 0.0))
            if current_share == 0:
                raise ValueError(f"Категория '{category}' отсутствует/нулевая доля в данных для '{var}'")
            adjustment = float(target_share) / current_share
            mask = df[var].astype(str) == str(category)
            weights.loc[mask] *= adjustment

    weights = weights.clip(TRIM_MIN, TRIM_MAX)
    if NORMALIZE_MEAN_1 and float(weights.mean()) != 0.0:
        weights = weights / float(weights.mean())

    df_out = df.copy()
    df_out["weight"] = weights.values
    return df_out


# =========================
# Derived variables (bot)
# =========================


def get_column_by_prefix(df: pd.DataFrame, prefix: str) -> str | None:
    """
    Ищет колонку по префиксу (как в telegram боте).
    Пример: "Q2" -> найдет "Q2 - ..." (и также "Q2 ..." / "Q2 - ...").
    """
    prefix_lower = str(prefix).lower().strip()
    for col in df.columns:
        col_lower = str(col).lower()
        if col_lower == prefix_lower:
            return col
        if col_lower.startswith(prefix_lower + " ") or col_lower.startswith(prefix_lower + " -"):
            return col
    return None


def _parse_int_list(text: str) -> List[int]:
    """
    Парсит список целых кодов: "1,2,3" -> [1,2,3]
    Диапазоны в этой функции НЕ обрабатываются (их обрабатываем отдельно для выражений).
    """
    parts = [p.strip() for p in (text or "").replace(";", ",").split(",") if p.strip()]
    out: List[int] = []
    for p in parts:
        out.append(int(float(p)))
    return out


def _apply_top_bot(df: pd.DataFrame, vars_list: List[str], op: str, suffix: str, custom_codes: List[int] | None = None) -> List[str]:
    """
    Создаёт бинарные переменные TOP/BOT.
    Для TOP/BOT берутся последние/первые N уникальных значений шкалы для каждой переменной отдельно.
    Для custom_codes — используются пользовательские коды для всех выбранных переменных.
    """
    created: List[str] = []
    if not vars_list:
        return created

    for var in vars_list:
        if var not in df.columns:
            continue
        series_num = pd.to_numeric(df[var], errors="coerce")
        unique_codes = sorted(series_num.dropna().astype(int).unique().tolist())

        codes: List[int]
        if custom_codes is not None:
            codes = custom_codes
        else:
            n = int(op.replace("top", "").replace("bot", "")) if op.lower().startswith(("top", "bot")) else 2
            # защита, если шкала слишком короткая
            n = max(1, n)
            if len(unique_codes) < n:
                codes = unique_codes
            else:
                if op.lower().startswith("top"):
                    codes = unique_codes[-n:]
                else:
                    codes = unique_codes[:n]

        new_var = f"{var}_{suffix}"
        df[new_var] = series_num.apply(lambda x: 1 if pd.notna(x) and int(x) in codes else 0).astype(int)
        created.append(new_var)

    return created


def _apply_mean(df: pd.DataFrame, vars_list: List[str]) -> List[str]:
    created: List[str] = []
    for var in vars_list:
        if var not in df.columns:
            continue
        series_num = pd.to_numeric(df[var], errors="coerce")
        mean_val = float(series_num.mean(skipna=True)) if len(series_num.dropna()) else 0.0
        new_var = f"{var}_MEAN"
        df[new_var] = mean_val
        created.append(new_var)
    return created


def _parse_group_mapping(text: str) -> Dict[int, List[int]]:
    """
    Ожидаемый формат (несколько вариантов):
      1:1,2,3; 2:4,5,6
      1=1,2,3 ; 2=4,5
    Возвращает {new_code: [old_code,...]}
    """
    raw = (text or "").strip()
    if not raw:
        raise ValueError("Пустое описание группировки кодов.")

    parts = [p.strip() for p in re.split(r"[;\n]+", raw) if p.strip()]
    if not parts:
        raise ValueError("Не удалось распознать группы в описании.")

    out: Dict[int, List[int]] = {}
    for part in parts:
        if ":" in part:
            left, right = part.split(":", 1)
        elif "=" in part:
            left, right = part.split("=", 1)
        else:
            raise ValueError(f"Неверный формат группы: '{part}'. Используйте 'new:old1,old2'.")
        new_code = int(float(left.strip()))
        old_codes = _parse_int_list(right)
        out[new_code] = old_codes
    return out


def _apply_group_codes(df: pd.DataFrame, var: str, mapping: Dict[int, List[int]], new_var_name: str) -> str:
    if var not in df.columns:
        raise ValueError(f"Переменная для группировки '{var}' не найдена в данных.")
    series_num = pd.to_numeric(df[var], errors="coerce")

    new_series = series_num.apply(lambda x: None if pd.isna(x) else int(x)).apply(
        lambda x: None if x is None else x
    )

    def map_code(x: Any) -> Any:
        if x is None or pd.isna(x):
            return None
        x_int = int(x)
        for new_code, old_codes in mapping.items():
            if x_int in old_codes:
                return new_code
        return None

    df[new_var_name] = new_series.apply(map_code)
    return new_var_name


def _evaluate_logic_expression(df: pd.DataFrame, expression: str) -> pd.Series:
    """
    Реплика evaluate_combined_expression из бота (с ограниченной безопасностью).
    Поддержка: =, <>, диапазоны в кодах (1-3), И/ИЛИ, скобки.
    """
    if not expression or not expression.strip():
        raise ValueError("Пустое выражение логики.")

    expr = str(expression).strip()
    expr = re.sub(r"\s*<>\s*", "<>", expr)
    expr = re.sub(r"\s*=\s*", "=", expr)
    expr = re.sub(r"\s*И\s*", " И ", expr, flags=re.I)
    expr = re.sub(r"\s*ИЛИ\s*", " ИЛИ ", expr, flags=re.I)
    expr = re.sub(r"\s+", " ", expr).strip()

    condition_pattern = r"([A-Za-z0-9_\-]+)(=|<>)([0-9,\-]+)"
    matches = re.findall(condition_pattern, expr)
    if not matches:
        raise ValueError("Не удалось распознать условия в выражении. Ожидается формат Q2=4,5 И Q3<>1-3.")

    condition_series: Dict[str, pd.Series] = {}
    expr_tmp = expr

    for i, (var_name, operator, codes_str) in enumerate(matches):
        full_var_name = var_name if var_name in df.columns else get_column_by_prefix(df, var_name)
        if not full_var_name:
            raise ValueError(f"Переменная '{var_name}' не найдена (и не удалось найти по префиксу).")

        # Парсинг кодов (диапазоны вида 1-3)
        codes: List[int]
        if "-" in codes_str and "," not in codes_str:
            start, end = map(int, codes_str.split("-"))
            codes = list(range(start, end + 1))
        else:
            codes = [int(c.strip()) for c in codes_str.split(",") if c.strip()]

        cond_key = f"{var_name}{operator}{codes_str}"
        cond_series = pd.to_numeric(df[full_var_name], errors="coerce").apply(
            lambda x: pd.notna(x) and int(x) in codes if operator == "=" else pd.notna(x) and int(x) not in codes
        )
        temp_name = f"__c{i}__"
        # заменяем именно восстановленную cond_key (spaces вокруг мы убрали выше)
        expr_tmp = expr_tmp.replace(cond_key, temp_name, 1)
        condition_series[temp_name] = cond_series.astype(bool)

    expr_python = expr_tmp.replace(" И ", " & ").replace(" ИЛИ ", " | ")
    # Подставляем в python выражение доступ только к condition_series
    for temp_name in list(condition_series.keys()):
        expr_python = expr_python.replace(temp_name, f"condition_series['{temp_name}']")

    # Быстрая проверка, что кроме condition_series[...] и операторов нет "лишних" токенов
    if re.search(r"[A-Za-z0-9_]", expr_python.replace("condition_series", "").replace("True", "").replace("False", "")):
        # это грубая проверка; если выражение было корректное, тут почти никогда не будет
        pass

    try:
        result_bool = eval(expr_python, {"__builtins__": {}}, {"condition_series": condition_series})
    except Exception as exc:
        raise ValueError(f"Ошибка вычисления выражения: {exc}")

    return result_bool.astype(int)


def _apply_derived_variables(df: pd.DataFrame, form: Any) -> Tuple[pd.DataFrame, List[str]]:
    """
    Применяет выбранные операции и возвращает (df_out, created_columns).
    """
    df_out = df.copy()
    created: List[str] = []

    # TOP/BOT
    topbot_vars = form.getlist("topbot_vars")
    if topbot_vars:
        selected_top2 = form.get("top2") is not None
        selected_top3 = form.get("top3") is not None
        selected_bot2 = form.get("bot2") is not None
        selected_bot3 = form.get("bot3") is not None
        custom_codes_enabled = form.get("topbot_custom_enabled") is not None
        custom_codes_text = form.get("topbot_custom_codes", "")

        if selected_top2:
            created.extend(_apply_top_bot(df_out, topbot_vars, op="top2", suffix="TOP2"))
        if selected_top3:
            created.extend(_apply_top_bot(df_out, topbot_vars, op="top3", suffix="TOP3"))
        if selected_bot2:
            created.extend(_apply_top_bot(df_out, topbot_vars, op="bot2", suffix="BOT2"))
        if selected_bot3:
            created.extend(_apply_top_bot(df_out, topbot_vars, op="bot3", suffix="BOT3"))

        if custom_codes_enabled:
            codes = _parse_int_list(custom_codes_text)
            created.extend(_apply_top_bot(df_out, topbot_vars, op="top_custom", suffix="TOP_CUSTOM", custom_codes=codes))

    # MEAN
    mean_vars = form.getlist("mean_vars")
    if mean_vars:
        created.extend(_apply_mean(df_out, mean_vars))

    # GROUP
    group_var = form.get("group_var")
    group_new_name = form.get("group_new_name")
    group_mapping_text = form.get("group_mapping")
    if group_var and group_mapping_text and group_new_name:
        mapping = _parse_group_mapping(group_mapping_text)
        created.append(_apply_group_codes(df_out, group_var, mapping, group_new_name))

    # LOGIC
    logic_expression = form.get("logic_expression", "")
    logic_new_name = form.get("logic_new_name", "")
    if logic_expression.strip() and logic_new_name.strip():
        logic_series = _evaluate_logic_expression(df_out, logic_expression)
        df_out[logic_new_name] = logic_series
        created.append(logic_new_name)

    # AUTO FREQ: TOP2/BOT2/MEAN for scales 1-5 / 1-9
    auto_freq_enabled = form.get("auto_freq_enabled") is not None
    if auto_freq_enabled:
        auto_candidates: List[str] = []
        for c in df_out.columns:
            if str(c).endswith(("_TOP2", "_TOP3", "_BOT2", "_BOT3", "_MEAN", "_TOP_CUSTOM")):
                continue
            s = pd.to_numeric(df_out[c], errors="coerce").dropna()
            if len(s) < 10:
                continue
            uniq = sorted(s.astype(int).unique().tolist())
            if len(uniq) < 4:
                continue
            # шкалы 1-5 или 1-9 (допускаем пропуски/частичное покрытие, но внутри диапазона)
            if min(uniq) >= 1 and max(uniq) <= 5:
                auto_candidates.append(c)
            elif min(uniq) >= 1 and max(uniq) <= 9:
                auto_candidates.append(c)

        # ограничим чтобы не раздуть файл
        auto_candidates = auto_candidates[:50]
        if auto_candidates:
            created.extend(_apply_top_bot(df_out, auto_candidates, op="top2", suffix="TOP2"))
            created.extend(_apply_top_bot(df_out, auto_candidates, op="bot2", suffix="BOT2"))
            created.extend(_apply_mean(df_out, auto_candidates))

    # Убираем служебные колонки, созданные нами ранее в процессе логики (если есть)
    return df_out, created


def cross_weight(df: pd.DataFrame, vars_list: List[str], targets: Dict[str, float]) -> pd.DataFrame:
    cross_var = create_cross_variable(df, vars_list)
    df2 = df.copy()
    df2["_cross_var"] = cross_var.astype(str)

    present = set(df2["_cross_var"].unique())
    declared = set(targets.keys())
    missing = declared - present
    if missing:
        raise ValueError(f"В данных отсутствуют категории: {', '.join(sorted(missing))}")

    weights = pd.Series(1.0, index=df2.index)
    current_shares = df2["_cross_var"].value_counts(normalize=True)

    for category, target_share in targets.items():
        current_share = float(current_shares.get(category, 0.0))
        if current_share == 0:
            raise ValueError(f"Категория '{category}' отсутствует в данных или имеет нулевую долю")
        adjustment = float(target_share) / current_share
        mask = df2["_cross_var"] == category
        weights.loc[mask] *= adjustment

    weights = weights.clip(TRIM_MIN, TRIM_MAX)
    if NORMALIZE_MEAN_1 and float(weights.mean()) != 0.0:
        weights = weights / float(weights.mean())

    cross_var_name = " × ".join(vars_list)
    df_out = df2.drop(columns=["_cross_var"], errors="ignore")
    df_out["weight"] = weights.values
    # Bot adds a cross variable name for diagnostics; it doesn't harm analysis.
    df_out[cross_var_name] = cross_var.astype(str)
    return df_out


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            file = request.files.get("data_file")
            if not file or not file.filename:
                flash("Пожалуйста, выберите файл с данными.", "error")
                return redirect(url_for("index"))

            try:
                df, all_values_map, col_labels = _load_dataframe_from_upload(file)
            except Exception as exc:
                flash(str(exc), "error")
                return redirect(url_for("index"))
            # Новый файл = новый источник данных. Старое очищаем.
            old_active = session.get("active_data_key")
            old_base = session.get("data_key")
            _drop_dataframe(old_active)
            if old_base and old_base != old_active:
                _drop_dataframe(old_base)
            _clear_outputs(session)

            try:
                key = _store_dataframe(df, all_values_map, col_labels)
                session["data_key"] = key
                session["active_data_key"] = key
                return redirect(url_for("variable_labels"))
            except Exception as exc:
                flash(f"Не удалось сохранить загруженные данные: {exc}", "error")
                return redirect(url_for("index"))
        except Exception as exc:
            traceback.print_exc()
            flash(f"Внутренняя ошибка при загрузке файла: {exc}", "error")
            return redirect(url_for("index"))

    return render_template("index.html")


@app.route("/variable-labels", methods=["GET", "POST"])
def variable_labels():
    """
    Шаг после загрузки базы: дерево переменных (код + метка вопроса), подгрузка datamap Excel.
    Не загружаем весь DataFrame из pickle (большие .sav) — только manifest колонок.
    """
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены, загрузите файл заново.", "error")
        return redirect(url_for("index"))
    cols = _read_column_manifest(key)
    if not cols:
        try:
            df_fb = _load_dataframe(key)
            cols = [str(c) for c in df_fb.columns]
        except Exception:
            flash("Не удалось загрузить данные из временного хранилища. Загрузите файл заново.", "error")
            return redirect(url_for("index"))

    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "continue":
            return redirect(url_for("configure"))
        if action == "clear_labels":
            COLUMN_LABELSTORE[key] = {}
            flash("Метки вопросов сброшены (имена столбцов в данных не менялись).", "info")
            return redirect(url_for("variable_labels"))
        if action == "upload_datamap":
            f = request.files.get("datamap_file")
            if not f or not f.filename:
                flash("Выберите файл datamap (Excel).", "error")
                return redirect(url_for("variable_labels"))
            ext = _upload_extension(f.filename)
            if ext not in {"xlsx", "xls"}:
                flash("Datamap: поддерживаются только Excel (.xlsx, .xls).", "error")
                return redirect(url_for("variable_labels"))
            suffix = ".xlsx" if ext == "xlsx" else ".xls"
            fd, tmp_path = tempfile.mkstemp(suffix=suffix)
            os.close(fd)
            mapping: Dict[str, str] = {}
            try:
                try:
                    f.stream.seek(0)
                except (OSError, AttributeError, ValueError):
                    pass
                f.save(tmp_path)
                mapping = _parse_datamap_excel(tmp_path)
            except Exception as exc:
                flash(f"Не удалось разобрать datamap: {exc}", "error")
                return redirect(url_for("variable_labels"))
            finally:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
            merge_mode = request.form.get("merge_mode", "overwrite")
            if merge_mode not in {"overwrite", "fill_empty"}:
                merge_mode = "overwrite"
            case_insensitive = bool(request.form.get("case_insensitive"))
            matched, updated = _apply_datamap_merge(key, cols, mapping, merge_mode, case_insensitive)
            return redirect(
                url_for(
                    "variable_labels",
                    success="datamap",
                    matched=matched,
                    updated=updated,
                    mode=merge_mode,
                )
            )

    datamap_success = request.args.get("success") == "datamap"
    datamap_matched = request.args.get("matched", type=int)
    datamap_updated = request.args.get("updated", type=int)
    datamap_merge_mode = (request.args.get("mode") or "").strip()
    datamap_applied_at = None
    if datamap_success:
        datamap_applied_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    try:
        items = _ui_column_items_from_columns(cols, key)
        groups = _ui_column_groups(items)
        n_labeled = sum(1 for x in items if x["has_label"] or x["header_is_descriptive"])
        return render_template(
            "variable_labels.html",
            n_columns=len(cols),
            column_items=items,
            column_groups=groups,
            columns=cols,
            n_column_labels=n_labeled,
            datamap_success=datamap_success,
            datamap_matched=datamap_matched,
            datamap_updated=datamap_updated,
            datamap_merge_mode=datamap_merge_mode,
            datamap_applied_at=datamap_applied_at,
        )
    except Exception as exc:
        # Защитный fallback: если страница дерева не смогла отрендериться
        # (например, слишком большой список переменных), продолжаем без падения 500.
        flash(
            "Не удалось отобразить дерево переменных после загрузки. "
            "Переходим сразу к настройке таблицы. "
            f"Детали: {exc}",
            "warning",
        )
        return redirect(url_for("configure"))


@app.route("/configure", methods=["GET", "POST"])
def configure():
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены, загрузите файл заново.", "error")
        return redirect(url_for("index"))

    try:
        df = _load_dataframe(key)
    except Exception:
        flash("Не удалось загрузить данные из временного хранилища. Загрузите файл заново.", "error")
        return redirect(url_for("index"))
    columns = list(df.columns)
    col_bundle = _template_column_bundle(df, key)
    closed_columns = _detect_closed_columns(df)

    if request.method == "POST":
        output_action = request.form.get("output_action", "show")  # show|download
        row_vars = request.form.getlist("row_vars")
        col_vars = request.form.getlist("col_vars")

        metric_count = bool(request.form.get("metric_count"))
        metric_percent = bool(request.form.get("metric_percent"))
        percent_base_total_sample = bool(request.form.get("percent_base_total_sample"))
        use_weights = bool(request.form.get("use_weights"))
        perform_ztest = bool(request.form.get("perform_ztest"))
        show_sig_marks = bool(request.form.get("show_sig_marks"))
        use_nested_columns = bool(request.form.get("use_nested_columns"))
        show_segment_coverage = bool(request.form.get("show_segment_coverage"))

        if not row_vars:
            flash("Выберите хотя бы одну переменную для строк.", "error")
            return redirect(url_for("configure"))

        if not col_vars:
            flash("Выберите хотя бы одну переменную для столбцов.", "error")
            return redirect(url_for("configure"))

        if not (metric_count or metric_percent):
            flash("Выберите хотя бы одну метрику.", "error")
            return redirect(url_for("configure"))

        config = SurveyConfig(
            row_vars=row_vars,
            col_vars=col_vars,
            metric_count=metric_count,
            metric_percent=metric_percent,
            percent_base_total_sample=percent_base_total_sample,
            use_weights=use_weights,
            perform_ztest=perform_ztest,
            show_sig_marks=show_sig_marks,
            use_nested_columns=use_nested_columns,
            show_segment_coverage=show_segment_coverage,
        )

        excel_id = None
        result_id = None
        try:
            table_df, significance = build_crosstab(
                df,
                config,
                OPTIONSTORE.get(key, {}),
                COLUMN_LABELSTORE.get(key, {}),
            )
        except Exception as exc:
            flash(str(exc), "error")
            return redirect(url_for("configure"))

        # Таблица для UI/Excel с явной маркировкой значимостей.
        display_df = _annotate_table_with_significance(table_df, significance) if config.show_sig_marks else table_df.copy()

        # В session НЕ храним config, т.к. список выбранных переменных может быть очень большим,
        # а Flask session хранится в cookie (JSON-сериализация может падать).

        # Excel: генерируем только если пользователь это запросил.
        if output_action == "download":
            try:
                cells = int(table_df.shape[0] * table_df.shape[1])
                if cells > 50000:
                    raise ValueError(
                        f"Excel слишком большой для выгрузки ({cells} ячеек). "
                        f"Уменьшите выбор переменных."
                    )
                excel_id = uuid.uuid4().hex
                tmp_name = f"/tmp/survey_web_excel_{excel_id}.xlsx"
                # В Excel сохраняем ЧИСТЫЕ значения (без стрелок), значимость только заливкой.
                table_df.to_excel(tmp_name, index=False)
                _apply_significance_fill_to_excel(tmp_name, table_df, significance)
                EXCELSTORE[excel_id] = tmp_name
                session["last_excel_id"] = excel_id
            except Exception as exc:
                session.pop("last_excel_id", None)
                flash(f"Не удалось подготовить Excel для скачивания: {exc}", "error")
                return redirect(url_for("results"))

            # Прямо скачиваем файл после подготовки
            return redirect(url_for("download_excel"))

        # show: только отображение (без Excel, чтобы не падать на to_excel)
        try:
            result_id = uuid.uuid4().hex
            table_html = display_df.to_html(classes="table table-sm table-striped", index=False)
            RESULTSTORE[result_id] = {
                "table_html": table_html,
                "significance": {
                    _significance_key_serialize(rq, ra, ck): v for (rq, ra, ck), v in significance.items()
                },
                "table_json": display_df.to_json(orient="split"),
                "table_raw_json": table_df.to_json(orient="split"),
                "config": config.__dict__,
            }
            session["last_result_id"] = result_id
        except Exception as exc:
            flash(f"Не удалось подготовить отображение таблицы: {exc}", "error")

        # Для режима "show" тоже подготовим Excel (но безопасно),
        # чтобы была возможность скачать после просмотра.
        try:
            cells = int(table_df.shape[0] * table_df.shape[1])
            if cells <= 50000:
                excel_id = uuid.uuid4().hex
                tmp_name = f"/tmp/survey_web_excel_{excel_id}.xlsx"
                table_df.to_excel(tmp_name, index=False)
                _apply_significance_fill_to_excel(tmp_name, table_df, significance)
                EXCELSTORE[excel_id] = tmp_name
                session["last_excel_id"] = excel_id
        except Exception:
            # Если Excel не подготовился - просто отключим кнопку скачивания
            session.pop("last_excel_id", None)

        return redirect(url_for("results"))

    has_weight = "weight" in columns

    # Prefill from last successful result to enable "retune" flow
    preselected_rows: List[str] = []
    preselected_cols: List[str] = []
    metric_count_default = True
    metric_percent_default = True
    percent_base_total_sample_default = False
    use_weights_default = has_weight
    perform_ztest_default = True
    show_sig_marks_default = True
    use_nested_columns_default = False
    show_segment_coverage_default = False

    if request.args.get("prefill") == "1":
        result_id = session.get("last_result_id")
        stored = RESULTSTORE.get(result_id, {}) if result_id else {}
        cfg = stored.get("config") if isinstance(stored, dict) else None
        if isinstance(cfg, dict):
            preselected_rows = [str(v) for v in cfg.get("row_vars", []) if str(v) in columns]
            preselected_cols = [str(v) for v in cfg.get("col_vars", []) if str(v) in columns]
            metric_count_default = bool(cfg.get("metric_count", True))
            metric_percent_default = bool(cfg.get("metric_percent", True))
            percent_base_total_sample_default = bool(cfg.get("percent_base_total_sample", False))
            use_weights_default = bool(cfg.get("use_weights", use_weights_default)) and has_weight
            perform_ztest_default = bool(cfg.get("perform_ztest", True))
            show_sig_marks_default = bool(cfg.get("show_sig_marks", True))
            use_nested_columns_default = bool(cfg.get("use_nested_columns", False))
            show_segment_coverage_default = bool(cfg.get("show_segment_coverage", False))

    return render_template(
        "configure.html",
        closed_columns=closed_columns,
        has_weight=has_weight,
        preselected_rows=preselected_rows,
        preselected_cols=preselected_cols,
        metric_count_default=metric_count_default,
        metric_percent_default=metric_percent_default,
        percent_base_total_sample_default=percent_base_total_sample_default,
        use_weights_default=use_weights_default,
        perform_ztest_default=perform_ztest_default,
        show_sig_marks_default=show_sig_marks_default,
        use_nested_columns_default=use_nested_columns_default,
        show_segment_coverage_default=show_segment_coverage_default,
        **col_bundle,
    )


@app.route("/results")
def results():
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены, загрузите файл заново.", "error")
        return redirect(url_for("index"))
    try:
        df = _load_dataframe(key)
    except Exception:
        flash("Не удалось загрузить данные из временного хранилища. Загрузите файл заново.", "error")
        return redirect(url_for("index"))
    columns = list(df.columns)
    col_bundle = _template_column_bundle(df, key)
    closed_columns = _detect_closed_columns(df)

    result_id = session.get("last_result_id")
    stored = RESULTSTORE.get(result_id, {}) if result_id else {}
    table_html = stored.get("table_html")
    significance_raw = stored.get("significance", {})
    cfg = stored.get("config", {}) if isinstance(stored, dict) else {}
    significance = _significance_dict_from_storage(
        significance_raw if isinstance(significance_raw, dict) else {}
    )

    sig_total = len(significance_raw) if isinstance(significance_raw, dict) else 0
    sig_significant = (
        sum(1 for _k, v in significance_raw.items() if str(v.get("direction")) != "none")
        if isinstance(significance_raw, dict)
        else 0
    )

    excel_id = session.get("last_excel_id")
    excel_ready = bool(excel_id and excel_id in EXCELSTORE and os.path.exists(EXCELSTORE.get(excel_id, "")))

    # defaults for quick retune panel on results page
    preselected_rows = [str(v) for v in cfg.get("row_vars", []) if str(v) in columns] if isinstance(cfg, dict) else []
    preselected_cols = [str(v) for v in cfg.get("col_vars", []) if str(v) in columns] if isinstance(cfg, dict) else []
    metric_count_default = bool(cfg.get("metric_count", True)) if isinstance(cfg, dict) else True
    metric_percent_default = bool(cfg.get("metric_percent", True)) if isinstance(cfg, dict) else True
    percent_base_total_sample_default = bool(cfg.get("percent_base_total_sample", False)) if isinstance(cfg, dict) else False
    has_weight = "weight" in columns
    use_weights_default = bool(cfg.get("use_weights", has_weight)) if isinstance(cfg, dict) else has_weight
    if not has_weight:
        use_weights_default = False
    perform_ztest_default = bool(cfg.get("perform_ztest", True)) if isinstance(cfg, dict) else True
    show_sig_marks_default = bool(cfg.get("show_sig_marks", True)) if isinstance(cfg, dict) else True
    use_nested_columns_default = bool(cfg.get("use_nested_columns", False)) if isinstance(cfg, dict) else False
    show_segment_coverage_default = bool(cfg.get("show_segment_coverage", False)) if isinstance(cfg, dict) else False

    return render_template(
        "results.html",
        table_html=table_html,
        significance=significance,
        excel_ready=excel_ready,
        sig_total=sig_total,
        sig_significant=sig_significant,
        last_config=cfg if isinstance(cfg, dict) else {},
        closed_columns=closed_columns,
        has_weight=has_weight,
        preselected_rows=preselected_rows,
        preselected_cols=preselected_cols,
        metric_count_default=metric_count_default,
        metric_percent_default=metric_percent_default,
        percent_base_total_sample_default=percent_base_total_sample_default,
        use_weights_default=use_weights_default,
        perform_ztest_default=perform_ztest_default,
        show_sig_marks_default=show_sig_marks_default,
        use_nested_columns_default=use_nested_columns_default,
        show_segment_coverage_default=show_segment_coverage_default,
        **col_bundle,
    )


@app.route("/download-excel")
def download_excel():
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены, загрузите файл заново.", "error")
        return redirect(url_for("index"))

    excel_id = session.get("last_excel_id")
    if excel_id and excel_id in EXCELSTORE and os.path.exists(EXCELSTORE.get(excel_id, "")):
        tmp_name = EXCELSTORE[excel_id]

        @after_this_request
        def _cleanup(response):  # noqa: ANN001
            try:
                os.remove(tmp_name)
            except OSError:
                pass
            EXCELSTORE.pop(excel_id, None)
            session.pop("last_excel_id", None)
            return response

        return send_file(
            tmp_name,
            as_attachment=True,
            download_name="survey_analysis.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Fallback: если временный excel уже удалился, соберем файл из сохраненного результата.
    result_id = session.get("last_result_id")
    stored = RESULTSTORE.get(result_id, {}) if result_id else {}
    table_json = stored.get("table_json")
    table_raw_json = stored.get("table_raw_json")
    significance_raw = stored.get("significance", {})
    significance = _significance_dict_from_storage(
        significance_raw if isinstance(significance_raw, dict) else {}
    )
    if table_raw_json or table_json:
        try:
            # Для Excel используем raw-таблицу без стрелок и добавляем только заливку.
            src_json = table_raw_json if table_raw_json else table_json
            table_df = pd.read_json(src_json, orient="split")
            tmp_name = f"/tmp/survey_web_excel_{uuid.uuid4().hex}.xlsx"
            table_df.to_excel(tmp_name, index=False)
            _apply_significance_fill_to_excel(tmp_name, table_df, significance)
            return send_file(
                tmp_name,
                as_attachment=True,
                download_name="survey_analysis.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            flash(f"Не удалось сформировать Excel из результата: {exc}", "error")
            return redirect(url_for("results"))

    flash("Excel пока не подготовлен. Нажмите 'Скачать Excel сразу' на странице расчёта.", "error")
    return redirect(url_for("results"))


@app.route("/weight", methods=["GET", "POST"])
def weight_page():
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены, загрузите файл заново.", "error")
        return redirect(url_for("index"))

    try:
        df = _load_dataframe(key)
    except Exception:
        flash("Не удалось загрузить данные из временного хранилища. Загрузите файл заново.", "error")
        return redirect(url_for("index"))
    columns = list(df.columns)
    col_bundle = _template_column_bundle(df, key)

    if request.method == "POST":
        weight_vars = request.form.getlist("weight_vars")
        weighting_type = request.form.get("weighting_type", "cross")
        action = request.form.get("action", "weigh")

        if not weight_vars:
            flash("Выберите хотя бы одну переменную для взвешивания.", "error")
            return redirect(url_for("weight_page"))

        if len(weight_vars) > MAX_VARS:
            flash(f"Можно выбрать не более {MAX_VARS} переменных.", "error")
            return redirect(url_for("weight_page"))

        file_targets = request.files.get("targets_file")
        targets_text = request.form.get("targets_text", "")

        if weighting_type == "cross":
            cross_var = create_cross_variable(df, weight_vars)
            categories = sorted(cross_var.dropna().unique().tolist())
            if len(categories) > MAX_CATS_PER_DIM:
                flash(
                    f"Слишком много категорий ({len(categories)}) для кросс‑взвешивания. Максимум: {MAX_CATS_PER_DIM}.",
                    "error",
                )
                return redirect(url_for("weight_page"))
            current_shares = cross_var.value_counts(normalize=True).sort_index()

            if action == "template":
                out_path = f"/tmp/targets_template_{uuid.uuid4().hex}.xlsx"
                create_targets_template_xlsx(categories, current_shares, out_path)
                return send_file(
                    out_path,
                    as_attachment=True,
                    download_name="targets_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # action == "weigh"
            targets: Dict[str, float] = {}
            try:
                if file_targets and file_targets.filename:
                    targets = read_targets_from_excel(file_targets)
                    missing = set(categories) - set(targets.keys())
                    extra = set(targets.keys()) - set(categories)
                    if missing or extra:
                        msg = []
                        if missing:
                            msg.append("нет значений для: " + ", ".join(sorted(missing)))
                        if extra:
                            msg.append("лишние категории: " + ", ".join(sorted(extra)))
                        raise ValueError("; ".join(msg))
                else:
                    targets = parse_targets_input(targets_text, categories)
            except Exception as exc:
                flash(f"Ошибка с целевыми значениями: {exc}", "error")
                return redirect(url_for("weight_page"))

            try:
                df_weighted = cross_weight(df, weight_vars, targets)
            except Exception as exc:
                flash(f"Ошибка взвешивания: {exc}", "error")
                return redirect(url_for("weight_page"))

        elif weighting_type == "sequential":
            # Ограничиваем размер категорий каждой переменной (чтобы шаблон был не слишком огромным)
            for var in weight_vars:
                cats = df[var].dropna().astype(str).unique()
                if len(cats) > MAX_CATS_PER_DIM:
                    flash(
                        f"Слишком много категорий ({len(cats)}) для переменной '{var}'. Максимум: {MAX_CATS_PER_DIM}.",
                        "error",
                    )
                    return redirect(url_for("weight_page"))

            if action == "template":
                out_path = f"/tmp/sequential_targets_template_{uuid.uuid4().hex}.xlsx"
                create_sequential_targets_template_xlsx(df, weight_vars, out_path)
                return send_file(
                    out_path,
                    as_attachment=True,
                    download_name="sequential_targets_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # action == "weigh"
            if not (file_targets and file_targets.filename):
                flash("Для sequential загрузите Excel с targets_file.", "error")
                return redirect(url_for("weight_page"))

            try:
                targets_by_var = read_sequential_targets_from_excel(file_targets, weight_vars)
                df_weighted = sequential_weight(df, weight_vars, targets_by_var)
            except Exception as exc:
                flash(f"Ошибка взвешивания: {exc}", "error")
                return redirect(url_for("weight_page"))

        else:
            flash("Неизвестный тип взвешивания.", "error")
            return redirect(url_for("weight_page"))

        _replace_active_dataframe(session, df_weighted)

        pretty_type = "кросс‑взвешивание" if weighting_type == "cross" else "последовательное взвешивание"
        flash(f"Взвешивание завершено ({pretty_type}). Теперь можно посчитать кросс‑таблицы.", "success")
        return redirect(url_for("configure"))

    return render_template("weight.html", **col_bundle)


@app.route("/variables", methods=["GET", "POST"])
def variables_page():
    key = session.get("active_data_key") or session.get("data_key")
    if not key or key not in DATASTORE:
        flash("Данные не найдены. Загрузите файл заново.", "error")
        return redirect(url_for("index"))

    try:
        df = _load_dataframe(key)
    except Exception:
        flash("Не удалось загрузить данные из временного хранилища. Загрузите файл заново.", "error")
        return redirect(url_for("index"))
    columns = list(df.columns)
    col_bundle = _template_column_bundle(df, key)
    closed_columns = _detect_closed_columns(df)

    allowed_return = {"/configure", "/weight", "/variable-labels"}
    return_to = request.args.get("return_to", "/configure")
    if return_to not in allowed_return:
        return_to = "/configure"

    if request.method == "POST":
        # Жёсткие лимиты на количество выбранных переменных для безопасности по памяти
        topbot_vars = request.form.getlist("topbot_vars")
        mean_vars = request.form.getlist("mean_vars")
        if len(topbot_vars) > 25:
            flash("TOP/BOT: выбери не более 25 переменных.", "error")
            return redirect(url_for("variables_page", return_to=return_to))
        if len(mean_vars) > 25:
            flash("MEAN: выбери не более 25 переменных.", "error")
            return redirect(url_for("variables_page", return_to=return_to))

        try:
            df_out, created = _apply_derived_variables(df, request.form)
        except Exception as exc:
            flash(f"Ошибка создания переменных: {exc}", "error")
            return redirect(url_for("variables_page", return_to=return_to))

        # Flask session хранится в cookie и сериализуется в JSON.
        # Приводим к plain string, чтобы избежать "not JSON serializable" ошибок.
        created = [str(x) for x in (created or [])]

        _replace_active_dataframe(session, df_out)
        # Храним только небольшой список, иначе session cookie может стать слишком большой.
        session["last_created_vars"] = created[:100]
        session["variables_return_to"] = return_to

        flash(
            f"Переменные созданы: {', '.join(created[:10])}{'...' if len(created) > 10 else ''}",
            "success",
        )
        return redirect(url_for("variables_page", return_to=return_to, show_result=1))

    show_result = request.args.get("show_result") == "1"
    created = session.get("last_created_vars", [])
    return render_template(
        "variables.html",
        closed_columns=closed_columns,
        return_to=return_to,
        show_result=show_result,
        created=created,
        **col_bundle,
    )


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=True)

